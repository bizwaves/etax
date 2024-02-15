# -*- coding: utf-8 -*-

###
# @file: eTaxAccCode.py
# @require: pip install openpyxl dataset requests
# @usage: eTaxAccCode.ps1 参照。（ご準備：Download Url 等は Table spec_file に登録済 ※ makeDocTable.py 参考）
# @version: 0.0.2, 2024.02.15
# @author: H. Nakano, nakano-h@bitwaves.org
# @url: https://www.bitwaves.org/
###

import pdb
import dataset
import pprint
import json
import requests
import pathlib
import sys
import traceback
from openpyxl import load_workbook, Workbook

from makeIndustryTable import Record_List as Industry_List

#=
# 勘定科目コード表の説明
#   （参考）１　EDINETとは、金融庁ホームページにおいて提供されている「金融商品取引法に基づく有価証券報告書等の開示書類に関する電子開示システム」である。									
#  　　　　 ２  上記にある各項目は、次のとおり引用又は当庁にて作成している。
# 　　　　  　　〔No.1～15〕　：　「2019年版EDINETタクソノミ対応『勘定科目リスト』」(EDINET閲覧サイト)を引用（注）
# 　　　　  　　〔No.0、16～18〕　：　「2019年版EDINETタクソノミ対応『勘定科目リスト』」(EDINET閲覧サイト)を基に当庁が作成
# 　　　　  　　〔No.19～22〕　：　新規作成設定
# 　　　  　　　（注）損益計算書の一部の勘定科目は、「No.12　period Type」を変更していますのでご留意ください。							
# 　　  　　３　〔No.6～14〕は「勘定科目コード表」では「非表示」設定にしていますので、閲覧される際は「再表示」にしてください。									
#=								
Field_Dict = {  # Row 2 : Row 3 ⇒ DB Field Name
    '0': {'name': 'industry',            'desc': '業種',                        'desc_en': 'Industry',                              "long_desc": "EDINETで設定されている業種目（23業種）を表示しています。"},
    "1": {'name': 'classification',      'desc': '科目分類',                    'desc_en': 'Classification',                        "long_desc": "財務諸表本表タクソノミに設定されている勘定科目の科目分類です。"},
    "2": {'name': 'std_label_ja',        'desc': '標準ラベル（日本語）',        'desc_en': 'Standard label (Japanese)',             "long_desc": "勘定科目の階層構造を標準ラベル（日本語）を用いて表しています。"},
    "3": {'name': 'ret_label_ja',        'desc': '冗長ラベル（日本語）',        'desc_en': 'Redundant label (Japanese)',            "long_desc": "勘定科目の冗長ラベル（日本語）です。"},
    "4": {'name': 'std_label_en',        'desc': '標準ラベル（英語）',          'desc_en': 'Standard label (English)',              "long_desc": "勘定科目の標準ラベル（英語）です。"},
    "5": {'name': 'ret_label_en',        'desc': '冗長ラベル（英語）',          'desc_en': 'Redundant label (English)',             "long_desc": "勘定科目の冗長ラベル（英語）です。"},
    "6": {'name': 'usage_cate_ja',       'desc': '用途区分（日本語）',          'desc_en': 'Usage category (Japanese)',             "long_desc": "勘定科目の用途区分、財務諸表区分及び業種区分のラベル（日本語）です。"},
    "7": {'name': 'usage_cate_en',       'desc': '用途区分（英語）',            'desc_en': 'Usage category (English)',              "long_desc": "勘定科目の用途区分、財務諸表区分及び業種区分のラベル（英語）です。"},
    "8": {'name': 'namespace_prefix',    'desc': '名前空間プレフィックス',      'desc_en': 'Namespace prefix',                      "long_desc": "勘定科目の名前空間プレフィックスです。"},
    "9": {'name': 'element_name',        'desc': '要素名',                      'desc_en': 'Element name',                          "long_desc": "勘定科目の要素名です。"},
    "10": {'name': 'type',               'desc': 'データ型',                    'desc_en': 'Data type',                             "long_desc": "勘定科目のデータ型（type属性）です。"},
    "11": {'name': 'substitution_group', 'desc': '代替グループ',                'desc_en': 'Substitution group',                    "long_desc": "勘定科目の代替グループ(substitutionGroup属性)です。"},
    "12": {'name': 'period_type',        'desc': '期間時点区分',                'desc_en': 'Period type',                           "long_desc": "勘定科目の期間時点区分（periodType属性）です。"},
    "13": {'name': 'balance',            'desc': '貸借区分',                    'desc_en': 'Balance',                               "long_desc": "勘定科目の貸借区分(balance属性)です。"},
    "14": {'name': 'abstract',           'desc': '抽象区分',                    'desc_en': 'Abstract',                              "long_desc": "勘定科目の抽象区分(abstract属性)です。"},
    "15": {'name': 'depth',               'desc': 'depth',                      'desc_en': 'Depth',                                 "long_desc": "勘定科目の科目一覧ツリー又はグローバルディメンションにおける階層情報です。"},
    "16": {'name': 'title_item',          'desc': 'タイトル項目',               'desc_en': 'Title item',                            "long_desc": "EDINETの勘定科目リストで冗長ラベルがタイトル項目である勘定科目には「○」で表示。"},
    "17": {'name': 'total_usage_cate',    'desc': '合計（用途区分）',           'desc_en': 'Total (usage category)',                "long_desc": "EDINETの勘定科目リストで用途区分が合計と使用できる勘定科目には「○」で表示。"},
    "18": {'name': 'acc_cate',            'desc': '勘定科目区分',               'desc_en': 'Account category',                      "long_desc": "EDINETの勘定科目リストで使用されている勘定科目を財務諸表規則に基づき区分したもの。"},
    "19": {'name': 'acc_code',            'desc': '勘定科目コード',             'desc_en': 'Account category code',                 "long_desc": "貸借対照表のCSV形式データを作成するのに使用する勘定科目コードです。"},
    "20": {'name': 'etax_acc_ja',         'desc': 'e-Tax対応勘定科目（日本語）', 'desc_en': 'e-Tax compatible account (Japanese)',  "long_desc": "EDINETの勘定科目リストで使用されている勘定科目及び勘定科目コードに対応する公表用e-Tax勘定科目（日本語）"},
    "21": {'name': 'etax_acc_en',         'desc': 'e-Tax対応勘定科目（英語）',  'desc_en': 'e-Tax compatible account (English)',    "long_desc": "EDINETの勘定科目リストで使用されている勘定科目及び勘定科目コードに対応する公表用のe-Tax勘定科目（英語）"},
    "22": {'name': 'etax_acc_cate',       'desc': 'e-Tax勘定科目区分',          'desc_en': 'e-Tax account category',                "long_desc": "e-Taxの勘定科目を財務諸表規則に基づき区分したもの。"},
}


def findSpecFile(db, name, variation, version):
    #=
    # DB Table spec_file からファイル名等を取得する 
    #=
    # pdb.set_trace()
    tbl_spec_file = db['spec_file']
    spec_file = tbl_spec_file.find_one(name = name, variation = variation, version = version)
    assert(spec_file is not None)
    pprint.pprint(spec_file)
    return spec_file


def makeDocEntryTable(db, name, variation, version, tableName, madeFlag):
    # = +=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=
    # Field_Dict を Doc Entry DB Table に upsert
    #
    # = +=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=
    # pdb.set_trace()

    spec_file = findSpecFile(db, name, variation, version)
    
    # データベース周りの準備
    table = db[tableName]

    # レコードをDB Tableに保存
    for k, v in Field_Dict.items():
        data = { "spec_name" : spec_file["name"] }  # Key Field 1
        assert(v["name"] == v["name"].lower())
        data["name"] = v["name"]                    # Key Field 2
        data["desc"] = v["desc"]
        data["desc_en"] = v["desc_en"]
        data["long_desc"] = v["long_desc"]
        data["made_flag"] = madeFlag    # 今回作成・修正に示すフラグ

        pprint.pprint(data)

        table.upsert(data, keys=[
            'spec_name', 'name'
        ])
    # rof

    print("makeDocEntryTable done")


def readXlsToTable(db, name, variation, version, tableName, madeFlag):
    # = +=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=
    # e-tax 勘定科目コード表 excel ファイルを読み込んで、Database Table に upsert する
    #   ※ ローカルファイルが存在しない場合は自動的ダウンロードする
    # = +=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=+=
    # pdb.set_trace()
    
    spec_file = findSpecFile(db, name, variation, version)
    
    localfile = spec_file['localfile']
    sheetName = spec_file['sheet']
    table = spec_file['table']
    

    assert(localfile is not None and len(localfile) > 0)
    assert(sheetName is not None and len(sheetName) > 0)
    assert(table is not None and len(table) > 0)
    
    
    xlsFile = pathlib.Path(localfile)
    
    # test parent  directory of local file exists or not
    if not xlsFile.parent.is_dir():
        print("Error: {} is not a directory or not exists.".format(xlsFile.parent.absolute()))
        exit(1)
    # fi

    if not xlsFile.is_file():
        # ファイルをダウンロードする
        curl = spec_file['curl']
        print(f"{xlsFile.name} not exists, Downloading {curl} to {xlsFile.absolute()}")
        assert(curl is not None and len(curl) > 0)
        response = requests.get(curl)
        if response.status_code == 200:
            with open(xlsFile, 'wb') as f:
                f.write(response.content)
        else:
            print(f"Error: Failed to download the file. Status code: {response.status_code}")
            # ↓ is_file check で exit しているので、ここで exit しない
        #fi
    else:
        print(f"{xlsFile.name} exists.")
    # fi

    if not xlsFile.is_file():
        print("Error: {} is not a file or not exists.".format(xlsFile.absolute()))
        exit(1)
    # fi
    
    # 拡張子チェック
    if not xlsFile.suffix.lower() in ['.xls', '.xlsx', '.xlsm']:
        print(
            "Error: The extension of the input file must be in ['.xls', '.xlsx', '.xlsm' ]")
        exit(1)
    # fi
    
    # Excel ファイルを読み込む
    wb = load_workbook(filename=xlsFile.as_posix(), data_only=True)
    ws = wb[sheetName]

    title = ws["A1"].value  # 貸借対照表　勘定科目コード表（2019年版）
    print(title)

    # ヘッダー行を読み出してフィールドリストを作成
    # ['0', '1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13', '14', '15', '16', '17', '18', '19', '20', '21', '22']
    headers = [cell.value for cell in ws[2]]
    print("headers: {0}".format(headers))

    desc_headers = [cell.value for cell in ws[3]]       # ['業種',
    print("desc_headers: {0}".format(desc_headers))

    # do some test
    assert (len(headers) == 23)
    assert (headers[0] == "0")
    assert (desc_headers[0] == '業種')

    # データベース周りの準備
    table = db[tableName]

    # Row 4 以降のデータ行を読み出し、レコードとしてDB Tableに保存
    for row in ws.iter_rows(min_row=4, values_only=True):
        record = dict(zip(headers, row))    # {'0': '一般商工業',

        # Field_Dict を適用し、DB 登録用レコードを作る
        data = {}
        for key in Field_Dict:
            if key in [ "16", "17" ]:
                data[Field_Dict[key]['name']] = record[key] is not None and record[key].strip() == "○"  # None も False としたことに注意
            elif key in [ "0" ]:
                #= ↓ Industry_List から industry_code を取得して登録する
                # Industry_List = [ { "industry_code" : "10",   "industry_name" : "一般商工業" }, ... ]
                #=
                assert(record[key] is not None)
                industry_name = record[key].strip()
                
                industry_name = industry_name.replace("（", "(").replace("）", ")") # 全角括弧を半角括弧に変換: "銀行・信託業（特定取引勘定設置銀行）" ⇒ "銀行・信託業(特定取引勘定設置銀行)" に統一する
                
                industry_code = None
                for industry in Industry_List:
                    if industry["industry_name"] == industry_name:
                        industry_code = industry["industry_code"]
                        break
                    # fi
                # rof
                
                if industry_code is None:
                    print(f"Error: industry_code not found for {industry_name}")
                    exit(1)
                #fi
                assert(industry_code is not None)
                
                data[Field_Dict[key]['name']] = industry_code
            elif key in [ "15" ]:  # depth, int
                assert(record[key] is not None)
                assert(type(record[key]) == int)
                data[Field_Dict[key]['name']] = record[key]
            else:
                data[Field_Dict[key]['name']] = record[key].strip() if record[key] is not None else ""
            # fi
        # for

        data["jis_code"] = ""           # 【ペンディング】jis-code table へのリンク ※検討中だが、とりあえず空欄を作っておく

        # 元出処（Unique Key Fields）
        data["spec_name"] = name
        data["spec_variation"] = variation
        data["spec_version"] = version
        
        data["made_flag"] = madeFlag    # 今回作成・修正に示すフラグ

        pprint.pprint(data)

        table.upsert(data, keys=[
            'industry', 'classification', 'std_label_ja', 'ret_label_ja', 'depth', 'title_item', 'account_category',  # 合っているか？
            'ac_cate_code',
            "spec_name", "spec_variation", "spec_version"
        ])
    # rof

    print("readXlsToTable done")


if __name__ == '__main__':
    # pdb.set_trace()
    argvs = sys.argv  # コマンドライン引数を格納したリストの取得
    argc = len(argvs)  # 引数の個数

    if argc != 7:
        print('Usage: \n python eTaxAccCode.py "doc_file.name" "doc_file.variation" "doc_file.version" "Database Connecting String" "Table Name" "Made Flag String"')
        print('Usage: \n python eTaxAccCode.py "bs.acc-cate-code.etax.acc" "all" "2019" "postgresql://postgres:postgres@localhost:25432/dskacc" "etax_account" "2014.02.13"')
        exit(1)
    # fi

    name = argvs[1]
    variation = argvs[2]
    version = argvs[3]
    dbConnStr = argvs[4]
    tableName = argvs[5]
    madeFlag = argvs[6]

    # データベースに接続
    try:
        db = dataset.connect(dbConnStr)
    except Exception as e:
        print("Error: {0}".format(e))
        exit(1)
    # yrt

    try:
        makeDocEntryTable(db, name, variation, version, 'doc_entry', madeFlag)
        readXlsToTable(db, name, variation, version, tableName, madeFlag)
    except Exception as e:
        traceback.print_exc()
    # yrt
    
    exit(0)
